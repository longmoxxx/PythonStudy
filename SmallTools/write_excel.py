# -*- coding: utf-8 -*-
# 读写2007 excel
import openpyxl
import os

def file_extension(file_path):
    return os.path.splitext(file_path)[1]

def get_file_name(file_path):
    return os.path.splitext(file_path)[0]

def writeExcel(file_path):
    excel_file_name = 'video.xlsx'
    wb = openpyxl.load_workbook(excel_file_name)
    i = 1
    sheet = wb.get_sheet_by_name("Sheet1")
    for root, dirs, files in os.walk(file_path):
        for f in files:
            if file_extension(f) == ".mp4" or file_extension(f) == ".mov"or file_extension(f) == ".mpg"or file_extension(f) == ".ts" \
                    or file_extension(f) == ".mpeg":
                fp = os.path.join(root, f)
                sheet.cell(row=i + 1, column=1, value=str(fp))#文件路径
                sheet.cell(row=i + 1, column=2, value=str(f)) #文件名称
                i=i+1
    wb.save(excel_file_name)
def Main():
    # write_name = r"D:\video"
    write_name = "F:\\video"
    writeExcel(write_name)
    pass


if __name__ == "__main__":
    Main()